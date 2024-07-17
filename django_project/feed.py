import os
import django
# Configure Django settings
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'django_project.settings')
django.setup()

#code
import openpyxl
from app_0.models import model_Hanja

#model_Person.objects.all().delete()

workbook = openpyxl.load_workbook('final.xlsx')
sheet1=workbook["Sheet1"]

for index, row in enumerate(sheet1.iter_rows(min_row=3, values_only=True), start=3):
    new_word=sheet1.cell(row=index, column=1).value
    new_one_stroke=sheet1.cell(row=index, column=2).value
    new_total_stroke=sheet1.cell(row=index, column=3).value
    new_meaning=sheet1.cell(row=index, column=4).value
    new_sound=sheet1.cell(row=index, column=5).value
    new_meaning_sound=sheet1.cell(row=index, column=6).value
    new_lev_read=sheet1.cell(row=index, column=7).value
    new_lev_write=sheet1.cell(row=index, column=8).value
    if new_lev_write is None: new_lev_write=''
    new_form=sheet1.cell(row=index, column=9).value
    if new_form is None: new_form=''

    new_explanation=sheet1.cell(row=index, column=10).value
    if new_explanation is None: new_explanation=''

    new_component_meaning=sheet1.cell(row=index, column=11).value

    new_component_sound=sheet1.cell(row=index, column=12).value
    if new_component_sound is None: new_component_sound=''
    new_trace=sheet1.cell(row=index, column=13).value
    new_origin=''
    if new_trace is None: 
        new_trace=''
    else: new_origin=new_trace[-1]
    
    try:
        new_Hanja = model_Hanja.objects.create( 
            word=new_word,
            one_stroke=new_one_stroke,
            total_stroke=new_total_stroke,
            meaning=new_meaning,
            sound=new_sound,
            meaning_sound=new_meaning_sound.strip(),
            lev_read=new_lev_read,
            lev_write=new_lev_write,
            form=new_form,
            explanation=new_explanation,
            component_meaning=new_component_meaning.strip(),
            component_sound=new_component_sound.strip(),
            trace=new_trace,
            origin=new_origin.strip()
        )
    except:
        try:
            new_Hanja = model_Hanja.objects.create( 
                word=new_word,
                one_stroke=new_one_stroke,
                total_stroke=new_total_stroke,
                meaning=new_meaning,
                sound=new_sound,
                meaning_sound=new_meaning_sound.strip(),
                lev_read=new_lev_read,
                lev_write=new_lev_write,
                form=new_form,
                explanation='',
                component_meaning=new_component_meaning.strip(),
                component_sound=new_component_sound.strip(),
                trace=new_trace,
                origin=new_origin.strip()
            )
        except: print(new_word)
print('sucess')

    
